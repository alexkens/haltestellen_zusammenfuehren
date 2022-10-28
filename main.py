import pandas as pd
import openpyxl as op




# 1. Haltestellen aus Datei: VGG_waiting_time_trips_full_v2.xlsx auslesen
# 2. für jede Haltestelle genau eine x,y GPS-Koordinate auslesen und abspeichern
# 3. Koordination in Datei: Endhaltestellen_Liste_alle_depots_v2.xlsx in Spalten geo_lat und  geo_long eintragen


# Constants
PATH = "/home/local/RL-INSTITUT/alexander.kens/Dokumente/RLI-Dokumente/Buffered/"

HALTESTELLEN_XLSX = "VGG_waiting_time_trips_full_v2.xlsx"
END_FILE_XLSX = "Endhaltestellen_Liste_alle_depots_v2.xlsx"
END_FILE_WORKSHEET = "Endhaltestellen_Liste_alle_depo"

COORD_LIST = [
    "grdfko.xlsx",
    "koord_bergstr.xlsx",
    "koord_darmstadt_dieburg.xlsx",
    "koord_odenwaldkreis.xlsx",
    "koord_rheinneckarkreis.xlsx",
]

DASH_9_TIMES = "---------"
END_FILE_GEO_LAT_C = "Unnamed: 16"
END_FILE_GEO_LONG_C = "Unnamed: 17"
END_FILE_ENDHALTESTELLE_C = "Unnamed: 1"

END_FILE_GEO_LAT = "Q"
END_FILE_GEO_LONG = "R"

# 1
def read_haltestellen():
    haltestellen = []

    # read hastellen_xlsx
    df = pd.read_excel(PATH + HALTESTELLEN_XLSX)

    # read haltestellen from column: arrival_name_ and put them in the list haltestellen
    for i in range(len(df)):
        haltestellen.append(df['arrival_name_'][i])

    return haltestellen


# 2
def read_coordinates(haltestellen, file_coordinates):
    coordinates = dict()        # dict with key=haltestelle, value=tuple(x,y)
    left_haltestellen = []

    # read haltestellen from list into dict
    for i in range(len(haltestellen)):
        coordinates[haltestellen[i]] = tuple((0, 0))

    # read file_coordinates
    df = pd.read_excel(PATH + file_coordinates)

    columns = df.columns
    counter = 0
    for column in columns:
        df.rename(columns={column: f'Column {counter}'}, inplace=True)
        counter += 1

    # loop through dict and look for the first GPS-cood and put it into the dict
    # column 3 = Haltestelle, x=column 5, y=column 6
    for haltestelle in coordinates:
        for i in range(len(df)):
            if df['Column 3'][i] == haltestelle:
                if df['Column 5'][i] != DASH_9_TIMES:
                    coordinates[haltestelle] = tuple((df['Column 5'][i], df['Column 6'][i]))
                    break
    for element in coordinates:
        if coordinates[element] == tuple((0, 0)):
            left_haltestellen.append(element)

    return coordinates, left_haltestellen


# 3
def save_coord_in_file(coordinates):

    # open END_FILE as a openpyxl.Workbook
    wb = op.load_workbook(filename=END_FILE_XLSX)

    # find the right openpyxl.Worksheet
    ws = wb[END_FILE_WORKSHEET]
    # print(ws['B3'].value)

    # loop through the column=Endhaltestelle and fill the columns geo_lat, geo_long with the coord
    for element in coordinates:
        x = coordinates[element][0]
        y = coordinates[element][1]
        for index in range(3, 185):
            cell_string = "B" + str(index)
            if ws[cell_string].value == element:
                ws[END_FILE_GEO_LAT + str(index)] = x
                ws[END_FILE_GEO_LONG + str(index)] = y

    wb.save(END_FILE_XLSX)


def update_coordinates(coordinates, coordinates_tmp):
    for element in coordinates_tmp:
        if element not in coordinates_tmp.keys():
            coordinates[element] = coordinates_tmp[element]
        else:
            coordinates[element] = coordinates_tmp[element]


def print_dict_or_list(data):
    if type(data) is dict:
        for e in data:
            print(e, " ", data[e])
    if type(data) is list:
        for i in range(len(data)):
            print(data[i])
    print()


if __name__ == '__main__':

    haltestellen = read_haltestellen()
    coordinates = dict()

    for i in range(5):
        coordinates_tmp, haltestellen = read_coordinates(haltestellen, COORD_LIST[i])
        update_coordinates(coordinates, coordinates_tmp)

    save_coord_in_file(coordinates)
    print("Fehlende Geo-Daten: ", haltestellen)


    """with open("coordinates_file", "w") as f:
            for element in coordinates:
                string = element+ ": " + "x=" + str(coordinates[element][0]) + ", y=" + str(coordinates[element][1]) + "\n"
                f.write(string)


    # path = "/home/local/RL-INSTITUT/alexander.kens/Dokumente/RLI-Dokumente/Buffered/Endhaltestellen_Liste_alle_depots_v2 Kopie.xlsx"
    """









